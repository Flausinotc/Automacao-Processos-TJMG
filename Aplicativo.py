# Importar as bibliotecas necessárias
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep
import openpyxl

# Definir o número da OAB
numero_oba = 133864

# Inicializar o driver do Chrome
driver = webdriver.Chrome()

# Acessar o site de consulta pública do TJMG
driver.get('https://pje-consulta-publica.tjmg.jus.br/')
sleep(3)

# Preencher o número da OAB e selecionar o estado
campo_oab = driver.find_element(By.XPATH,"//input[@id='fPP:Decoration:numeroOAB']")
campo_oab.send_keys(numero_oba)

campo_estado = driver.find_element(By.XPATH,"//select[@id='fPP:Decoration:estadoComboOAB']")
opcoes_estados = Select(campo_estado)
opcoes_estados.select_by_visible_text('SP')

# Clicar no botão de pesquisa
botaopesquisar = driver.find_element(By.XPATH,"//input[@id='fPP:searchProcessos']")
botaopesquisar.click()
sleep(7)

# Pegar links para os processos
processos_links = driver.find_elements(By.XPATH,"//b[@class='btn-block']")

# Iterar sobre os links dos processos
for processo_link in processos_links:
    processo_link.click()  # Clicar no link do processo
    sleep(5)
    janelas = driver.window_handles
    driver.switch_to.window(janelas[-1])  # Mudar para a nova janela/tab
    driver.set_window_size(1920,1080)  # Definir o tamanho da janela/tab

    # Extrair informações do processo
    numero_processo_element = driver.find_elements(By.XPATH,"//div[@class='col-sm-12 ']")[0]
    numero_processo = numero_processo_element.text

    data_distribuicao_element = driver.find_elements(By.XPATH,"//div[@class='value col-sm-12 ']")[1]
    data_distribuicao = data_distribuicao_element.text

    movimentacoes_elements = driver.find_elements(By.XPATH,"//div[@id='j_id132:processoEventoPanel_body']//tr[contains(@class,'rich-table-row')]//td//div//div//span")
    lista_movimentacoes = [movimentacao.text for movimentacao in movimentacoes_elements]

    # Abrir o arquivo Excel
    wordbook = openpyxl.load_workbook('dados.xlsx')

    try:
        # Verificar se a planilha já existe no arquivo Excel
        if numero_processo in wordbook.sheetnames:
            pagina_processo = wordbook[numero_processo]
        else:
            pagina_processo = wordbook.create_sheet(numero_processo)
            pagina_processo['A1'].value = "Numero do processo"
            pagina_processo['B1'].value = "Data Distribuição"
            pagina_processo['C1'].value = "Movimentações"

        # Preencher informações na planilha
        pagina_processo['A2'].value = numero_processo
        pagina_processo['B2'].value = data_distribuicao

        for index, linha in enumerate(pagina_processo.iter_rows(min_row=2, max_row=len(lista_movimentacoes), min_col=3, max_col=3)):
            for celula in linha:
                celula.value = lista_movimentacoes[index]

        # Salvar as alterações no arquivo Excel
        wordbook.save('dados.xlsx')
        driver.close()  # Fechar a janela/tab atual
        sleep(5)
        driver.switch_to.window(driver.window_handles[0])  # Voltar para a primeira janela/tab
    except Exception as error:
        print("Erro:", error)

# Fechar o driver do Chrome
driver.quit()
